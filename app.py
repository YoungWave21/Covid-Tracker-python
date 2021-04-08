import requests
import bs4
import tkinter as tk
from datetime import datetime, date
import openpyxl as xl

letters = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v',
           'w', 'x', 'y', 'z']

now = datetime.now()
current_time = now.strftime("%H:%M:%S")


def get_html_data(url):
    return requests.get(url)


# Retrieves the HTML page with all the divs
def get_html_soup(url):
    html_data = get_html_data(url)
    soup = bs4.BeautifulSoup(html_data.text, 'html.parser')
    return soup.find('div', class_="content-inner").find_all("div", id="maincounter-wrap")


# Basic formatting to set the width
def format_sheet(sheet):
    for i in letters:
        sheet.column_dimensions[i.capitalize()].width = 20
    sheet.cell(1, 2).value = "Time"


# reussable function to get the selected cell
def finds_cell(selected, sheet):
    cell = sheet.cell(selected, 2)
    while cell.value is not None:
        selected += 1
        cell = sheet.cell(selected, 2)
    return selected, cell


# This function gets the data for the world because the class names in the website is different and the URL works differently
def get_covid_data():
    # loading the worbook, has to be done each time the sheet needs to be changed
    wb = xl.load_workbook('Record.xlsx')
    sheet = wb['Sheet1']
    format_sheet(sheet)
    url = "https://worldometers.info/coronavirus/"
    info_div = get_html_soup(url)
    all_data = "World \n"
    selected = 1
    selected, time_cell = finds_cell(selected, sheet)
    step = 3
    for block in info_div:
        text = block.find("h1", class_=None).get_text()
        count = block.find("span", class_=None).get_text()
        # adding new lines to the previous data and sperating them into seperate lines
        all_data = f'{all_data}{text} {count} \n'
        sheet.cell(selected, step).value = count
        sheet.cell(1, step).value = text[0:-1]
        step += 1

    time_cell.value = str(date.today()) + " " + str(current_time)
    sheet.cell(selected, 1).value = "World"
    wb.save('Record.xlsx')
    return all_data


def get_country_data():
    # loading the worbook, has to be done each time the sheet needs to be changed
    wb = xl.load_workbook('Record.xlsx')
    sheet = wb['Sheet1']
    name = textfield.get()
    format_sheet(sheet)
    url = "https://worldometers.info/coronavirus/country/" + name
    info_div = get_html_soup(url)
    all_data = ""
    selected = 1
    selected, time_cell = finds_cell(selected, sheet)
    step = 3
    for block in info_div:
        text = str(block.find("h1", class_=None))[4:-5]
        count = str(block.find("span", class_=None))[6:-7]
        numbers = ""
        for word in list(count):
            if word.isdigit():
                numbers += word
        try:
            count = "{:,}".format(int(numbers))
        except ValueError:
            pass
        all_data = f'{all_data}{text} {count} \n'
        sheet.cell(selected, step).value = count
        step += 1

    time_cell.value = str(date.today()) + " " + str(current_time)
    sheet.cell(selected, 1).value = name.capitalize()
    wb.save('Record.xlsx')
    mainlabel['text'] = name.capitalize() + "\n" + all_data


def reload():
    new_data = get_covid_data()
    mainlabel['text'] = new_data


root = tk.Tk()
root.geometry("900x700")
root.title("Covid Tracker")
f = ("poppins", 25, "bold")

textfield = tk.Entry(root, width=50)
textfield.pack()

mainlabel = tk.Label(root, text=get_covid_data(), font=f)
mainlabel.pack()

gbtn = tk.Button(root, text="Get Data", font=f, relief='solid', command=get_country_data)
gbtn.pack()

rbtn = tk.Button(root, text="Reload", font=f, relief='solid', command=reload)
rbtn.pack()

root.mainloop()
